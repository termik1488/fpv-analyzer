import os
import sqlite3
from datetime import datetime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side,
    Alignment
)

DB_PATH = "db/fpv.db"


def export_to_excel(
    report_date=None
):

    conn = sqlite3.connect(DB_PATH)

    cursor = conn.cursor()


    if report_date:

        year, month, day = report_date.split("-")

        cursor.execute("""
            SELECT
                datetime,
                osd_name,
                flight_distance_km,
                video_freq,
                control_freq,
                control_bandwidth,
                protocol,
                lora_rate,
                spreading_factor,
                battery_start,
                battery_end,
                status,
                estimated_launch_distance_m
            FROM interceptions
            WHERE datetime LIKE ?
            ORDER BY datetime
        """, (f"{day}/{month}/{year}%",))

    else:

        cursor.execute("""
            SELECT
                datetime,
                osd_name,
                flight_distance_km,
                video_freq,
                control_freq,
                control_bandwidth,
                protocol,
                lora_rate,
                spreading_factor,
                battery_start,
                battery_end,
                status,
                estimated_launch_distance_m
            FROM interceptions
            ORDER BY datetime
        """)

    rows = cursor.fetchall()
    total_records = len(rows)

    suppressed_count = 0
    video_loss_count = 0
    strike_count = 0

    for row in rows:

        status = row[11]

        if status == "suppressed":
            suppressed_count += 1

        elif status == "video_loss":
            video_loss_count += 1

        elif status == "strike":
            strike_count += 1

    if total_records > 0:

        suppression_percent = round(
            suppressed_count * 100 / total_records,
            1
        )

    else:

        suppression_percent = 0

    unique_platforms = len(
        set(
            row[1]
            for row in rows
            if row[1]
        )
    )

    unique_video_freqs = len(
        set(
            row[3]
            for row in rows
            if row[3]
        )
    )
    conn.close()

    wb = Workbook()

    # ==================================================
    # INTERCEPTIONS
    # ==================================================

    ws = wb.active

    ws.title = "Interceptions"
    ws["A1"] = "Звіт по FPV перехопленням"

    ws["A2"] = (
        f"Дата звіту: {report_date}"
        if report_date
        else "Full Database Report"
    )

    ws["A3"] = (
        f"Сформовано: "
        f"{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )

    ws["F2"] = (
    f"Всього записів: "
    f"{total_records}"
    )

    ws["F3"] = (
        f"Унікальних платформ: "
        f"{unique_platforms}"
    )

    ws["F4"] = (
        f"Унікальних частот відео: "
        f"{unique_video_freqs}"
    )

    ws["H2"] = (
        f"Подавлено: {suppressed_count}"
    )

    ws["H3"] = (
        f"Втрата відеосигналу: {video_loss_count}"
    )

    ws["H4"] = (
        f"Нанесено уражень: {strike_count}"
    )

    ws["I2"] = (
        f"% подавлень: {suppression_percent}%"
    )

    summary_fill = PatternFill(
        start_color="D9EAD3",
        end_color="D9EAD3",
        fill_type="solid"
    )

    summary_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in range(2, 5):

        cell = ws[f"F{row}"]

        cell.fill = summary_fill

        cell.border = summary_border

        cell.alignment = Alignment(
            horizontal="left"
        )
    for row in range(2, 5):

        cell = ws[f"H{row}"]

        cell.fill = summary_fill
        cell.border = summary_border

        cell.alignment = Alignment(
            horizontal="left"
        )

    cell = ws["I2"]

    cell.fill = summary_fill
    cell.border = summary_border

    cell.alignment = Alignment(
        horizontal="left"
    )

    ws["F2"].font = Font(bold=True)
    ws["F3"].font = Font(bold=True)
    ws["F4"].font = Font(bold=True)

    ws["A1"].font = Font(
        bold=True,
        size=16
    )

    ws["A2"].font = Font(
        bold=True,
        size=12
    )

    ws["A3"].font = Font(
        italic=True
    )
    headers = [
        "Дата/Час",
        "Назва OSD",
        "Дальність польоту (км)",
        "Частота відео",
        "Частота керування",
        "Ширина каналу",
        "Протокол",
        "LoRa Rate",
        "SF",
        "Батарея старт (%)",
        "Батарея кінець (%)",
        "Статус",
        "Дистанція виявлення (м)"
    ]

    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid"
    )

        
    ws.freeze_panes = "A7"
    red_fill = PatternFill(
        start_color="FF9999",
        end_color="FF9999",
        fill_type="solid"
    )

    for row in rows:

        ws.append(row)

        battery_start = row[9]

        current_row = ws.max_row

        if (
            battery_start
             and
            battery_start < 800
        ):

            for cell in ws[current_row]:

                cell.fill = red_fill

    ws.auto_filter.ref = (
        f"A6:M{ws.max_row}"
    )

    gray_fill = PatternFill(
        start_color="F2F2F2",
        end_color="F2F2F2",
        fill_type="solid"
    )

    for sheet in wb.worksheets:

        for column_cells in sheet.columns:

            try:

                length = max(
                    len(str(cell.value))
                    if cell.value else 0
                    for cell in column_cells
                )

                sheet.column_dimensions[
                    column_cells[0].column_letter
                ].width = min(
                    length + 5,
                    50
                )

            except:
                pass

    # ==================================================
    # TOP CREWS
    # ==================================================

    crews_sheet = wb.create_sheet(
        "Top Crews"
    )

    crews_sheet.append([
        "Назва",
        "Кількість"
    ])

    crew_counter = Counter()

    for row in rows:

        name = row[1]

        if name:

            crew_counter[name] += 1

    for name, count in crew_counter.most_common():

        crews_sheet.append([
            name,
            count
        ])

    # ==================================================
    # VIDEO FREQUENCIES
    # ==================================================

    video_sheet = wb.create_sheet(
        "Video Frequencies"
    )

    video_sheet.append([
        "Частота",
        "Кількість"
    ])

    video_counter = Counter()

    for row in rows:

        freq = row[3]

        if freq:

            video_counter[freq] += 1

    for freq, count in video_counter.most_common():

        video_sheet.append([
            freq,
            count
        ])

# ==================================================
# VIDEO BANDS
# ==================================================

    band_counter = {
        "1.2 GHz": 0,
        "3.3 GHz": 0,
        "5.8 GHz": 0,
        "6-7.2 GHz": 0
    }

    for row in rows:

        freq = row[3]

        if not freq:
            continue
 
        if 900 <= freq < 2000:
            band_counter["1.2 GHz"] += 1

        elif 2900 <= freq < 4200:
            band_counter["3.3 GHz"] += 1

        elif 4200 <= freq < 6000:
            band_counter["5.8 GHz"] += 1

        elif 6000 <= freq <= 7200:
            band_counter["6-7.2 GHz"] += 1

    video_sheet["E1"] = "Діапазон"
    video_sheet["F1"] = "Кількість"
    video_sheet["G1"] = "Відсоток"
    video_sheet["H1"] = "Графік"

    total_bands = sum(
        band_counter.values()
    )

    row_num = 2

    for band, count in band_counter.items():

        if total_bands > 0:

            percent = round(
                count * 100 / total_bands,
                1
            )

        else:

            percent = 0
  
        bars = "█" * round(
            percent / 2
        )

        video_sheet.cell(
            row=row_num,
            column=5,
            value=band
        )
  
        video_sheet.cell(
            row=row_num,
            column=6,
            value=count
        )

        video_sheet.cell(
            row=row_num,
            column=7,
            value=f"{percent}%"
        )

        video_sheet.cell(
            row=row_num,
            column=8,
            value=bars
        )

        row_num += 1

    # ==================================================
    # CONTROL RANGES
    # ==================================================

    control_sheet = wb.create_sheet(
        "Control Ranges"
    )

    control_sheet.append([
        "Діапазон",
        "Кількість",
        "Відсоток",
        "Графік"
    ])

    ranges = {
        "<300 MHz": 0,
        "300-500 MHz": 0,
        "500-700 MHz": 0,
        "700-1000 MHz": 0,
        "1000-1500 MHz": 0,
        "1500-2000 MHz": 0,
        "2000-2400 MHz": 0,
        "2400-2700 MHz": 0,
        ">2700 MHz": 0
    }

    for row in rows:

        freq = row[4]

        if freq is None:
            continue

        if freq < 300:

            ranges["<300 MHz"] += 1

        elif freq < 500:

            ranges["300-500 MHz"] += 1

        elif freq < 700:

            ranges["500-700 MHz"] += 1

        elif freq < 1000:

            ranges["700-1000 MHz"] += 1

        elif freq < 1500:

            ranges["1000-1500 MHz"] += 1

        elif freq < 2000:

            ranges["1500-2000 MHz"] += 1

        elif freq < 2400:

            ranges["2000-2400 MHz"] += 1

        elif freq <= 2700:

            ranges["2400-2700 MHz"] += 1

        else:

            ranges[">2700 MHz"] += 1

    total = sum(
        ranges.values()
    )

    for name, count in ranges.items():

        if total > 0:

            percent = round(
                count * 100 / total,
                1
            )

        else:

            percent = 0

        bars = "█" * round(percent / 2)

        control_sheet.append([
            name,
            count,
            f"{percent}%",
            bars
        ])

    control_sheet.append([
        "TOTAL",
        total,
        "100%"
    ])

    # ==================================================
    # PLATFORM STATS
    # ==================================================

    platform_sheet = wb.create_sheet(
        "Platform Stats"
    )

    platform_sheet.append([
        "Платформа",
        "Вильотів",
        "Середня дальність",
        "Макс. дальність"
    ])

    KNOWN_PLATFORMS = {
        "МОЛНІЯ",
        "SUDNY DEN",
        "VT40",
        "VT 40",
        "UT 40",
        "SIMARGL",
        "UT 40 SIMARGL",
        "MALOY VT 40"
    }

    platform_stats = {}

    for row in rows:

        name = row[1]
        distance = row[2]

        if name not in KNOWN_PLATFORMS:
            continue

        if distance is None:
            continue

        platform_stats.setdefault(
            name,
            []
        ).append(distance)

    for name, distances in sorted(
        platform_stats.items(),
        key=lambda x: len(x[1]),
        reverse=True
    ):

        platform_sheet.append([
            name,
            len(distances),
            round(
                sum(distances) / len(distances),
                2
            ),
            round(
                max(distances),
                2
            )
        ])

    # ==================================================
    # SAVE
    # ==================================================

    os.makedirs(
        "data/reports",
        exist_ok=True
    )

    if report_date:

        output_file = (
            f"data/reports/"
            f"FPV_Report_{report_date}.xlsx"
        )

    else:

        output_file = (
           "data/reports/"
            "FPV_Full_Report.xlsx"
        )


    ws.freeze_panes = "A7"

    video_sheet.freeze_panes = "A2"

    control_sheet.freeze_panes = "A2"

    platform_sheet.freeze_panes = "A2"

    crews_sheet.freeze_panes = "A2"
# ОФОРМЛЕННЯ ШАПОК

    thick = Side(
        border_style="medium",
        color="000000"
    )

    for sheet in wb.worksheets:

        header_row = 1

        if sheet.title == "Interceptions":
            header_row = 6

        for cell in sheet[header_row]:

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = header_fill

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            cell.border = Border(
                left=thick,
                right=thick,
                top=thick,
                bottom=thick
            )

    gray_fill = PatternFill(
        start_color="F2F2F2",
        end_color="F2F2F2",
        fill_type="solid"
    )

    for sheet in wb.worksheets:

        start_row = 2

        if sheet.title == "Interceptions":
            start_row = 7

        for row_num in range(
            start_row,
            sheet.max_row + 1
        ):

            if row_num % 2 == 0:

                for cell in sheet[row_num]:

                    if cell.fill.fill_type is None:

                        cell.fill = gray_fill

    for row_num in range(7, ws.max_row + 1):

        if row_num % 2 == 0:

            for cell in ws[row_num]:

                if cell.fill.fill_type is None:

                    cell.fill = gray_fill

    for sheet in wb.worksheets:

        for column_cells in sheet.columns:

            try:

                length = max(
                    len(str(cell.value))
                    if cell.value else 0
                    for cell in column_cells
                )

                sheet.column_dimensions[
                    column_cells[0].column_letter
                ].width = min(
                    max(length + 5, 15),
                    50
                )

            except:
                pass

    wb.save(output_file)

    print(
        f"Excel exported: {output_file}"
    )

    return output_file